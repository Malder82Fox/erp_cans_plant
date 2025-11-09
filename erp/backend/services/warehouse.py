"""Warehouse service layer."""
from __future__ import annotations

import csv
import io
from dataclasses import asdict, dataclass
from decimal import Decimal, InvalidOperation
from typing import Iterable, Optional

from fastapi import HTTPException, UploadFile, status
from sqlalchemy.orm import Session

from erp.backend.models.warehouse import AuditLog, Part
from erp.backend.repositories.warehouse import AuditLogRepository, PartRepository
from erp.backend.schemas.warehouse import ImportResult, PartCreate, PartUpdate


@dataclass
class ImportStatistics:
    """Aggregate counters for import results."""

    created: int = 0
    skipped: int = 0
    errors: int = 0

    def to_result(self) -> ImportResult:
        return ImportResult(**asdict(self))


class WarehouseService:
    """Service orchestrating warehouse operations."""

    def __init__(self, session: Session):
        self.parts = PartRepository(session)
        self.audit = AuditLogRepository(session)
        self.session = session

    def list_parts(
        self,
        q: Optional[str],
        category_id: Optional[int],
        location_id: Optional[int],
        vendor_id: Optional[int],
        low_stock: bool,
        sort_field: str,
        sort_dir: str,
    ):
        return self.parts.search(q, category_id, location_id, vendor_id, low_stock, sort_field, sort_dir)

    def get_part(self, part_id: int) -> Part:
        part = self.parts.get_by_id(part_id)
        if not part or part.is_deleted:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Part not found")
        return part

    def create_part(self, payload: PartCreate, user_id: int | None) -> Part:
        if self.parts.get_by_code(payload.part_code):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Part code already exists")
        part = Part(**payload.model_dump())
        self.parts.create(part)
        self.audit.create("Part", part.id, "create", user_id, changes=payload.model_dump_json())
        return part

    def update_part(self, part_id: int, payload: PartUpdate, user_id: int | None) -> Part:
        part = self.get_part(part_id)
        update_data = payload.model_dump(exclude_unset=True)
        for key, value in update_data.items():
            setattr(part, key, value)
        self.audit.create("Part", part.id, "update", user_id, changes=payload.model_dump_json(exclude_unset=True))
        return part

    def delete_part(self, part_id: int, user_id: int | None) -> None:
        part = self.get_part(part_id)
        self.parts.soft_delete(part)
        self.audit.create("Part", part.id, "delete", user_id, changes=None)

    def import_parts(self, upload: UploadFile, mapping: dict[str, str]) -> ImportResult:
        stats = ImportStatistics()
        content = upload.file.read()
        if upload.filename and upload.filename.lower().endswith(".csv"):
            reader = csv.DictReader(io.StringIO(content.decode("utf-8")))
            for row in reader:
                self._process_import_row(row, mapping, stats)
        elif upload.filename and upload.filename.lower().endswith((".xlsx", ".xls")):
            try:
                from openpyxl import load_workbook
            except ImportError as exc:  # pragma: no cover - optional dependency
                raise HTTPException(status_code=500, detail="openpyxl required for XLSX import") from exc
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            header_row = next(sheet.iter_rows(values_only=True))
            headers = [str(value) if value is not None else "" for value in header_row]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                record = {header: value for header, value in zip(headers, row)}
                self._process_import_row(record, mapping, stats)
        else:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported file format")
        return stats.to_result()

    def _process_import_row(self, row: dict[str, str], mapping: dict[str, str], stats: ImportStatistics) -> None:
        try:
            part_code_key = mapping.get("part_code", "part_code")
            part_code = str(row.get(part_code_key, "")).strip()
            if not part_code:
                stats.skipped += 1
                return
            if self.parts.get_by_code(part_code):
                stats.skipped += 1
                return
            qty_key = mapping.get("qty_on_hand", "qty_on_hand")
            min_stock_key = mapping.get("min_stock", "min_stock")
            price_key = mapping.get("price", "price")
            qty_on_hand = self._parse_decimal(row.get(qty_key))
            min_stock = self._parse_decimal(row.get(min_stock_key))
            price = self._parse_decimal(row.get(price_key))
            data = {
                "part_code": part_code,
                "name": str(row.get(mapping.get("name", "name"), part_code)).strip(),
                "description": row.get(mapping.get("description", "description")),
                "qty_on_hand": qty_on_hand,
                "min_stock": min_stock,
                "price": price,
                "currency": str(row.get(mapping.get("currency", "currency"), "USD"))[:3],
            }
            part = Part(**data)
            self.parts.create(part)
            stats.created += 1
            self.audit.create("Part", part.id, "import", None, changes=str(data))
        except (InvalidOperation, KeyError, TypeError, ValueError):
            stats.errors += 1
        except Exception:  # noqa: BLE001
            stats.errors += 1

    def export_parts(self, parts: Iterable[Part]) -> str:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "id",
            "part_code",
            "name",
            "qty_on_hand",
            "min_stock",
            "price",
            "currency",
        ])
        for part in parts:
            writer.writerow([
                part.id,
                part.part_code,
                part.name,
                str(part.qty_on_hand),
                str(part.min_stock),
                str(part.price),
                part.currency,
            ])
        return output.getvalue()

    def list_audit_logs(self, part_id: int) -> Iterable[AuditLog]:
        return self.audit.list_for_entity("Part", part_id)

    @staticmethod
    def _parse_decimal(value: object) -> Decimal:
        if value in (None, ""):
            return Decimal("0")
        return Decimal(str(value))
